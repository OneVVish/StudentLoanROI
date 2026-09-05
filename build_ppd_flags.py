#!/usr/bin/env python3
"""Which programmes PPD:2026 says would fail the OBBBA earnings test.

Reads the Department of Education's Program Performance Data (PPD:2026) and
writes one row per (school, programme) that the app can look up, fanned out
from the Title IV certification to the campuses under it through
data/ppd_opeid_crosswalk.csv.

WHAT PPD IS. Released 2025-12-30 ahead of the AHEAD negotiated rulemaking,
209,321 programmes at 5,096 institutions. ED states plainly that it is NOT the
official eligibility metric, that these "are not the final data that will be
used to implement the proposed rule", and that the public file "differs somewhat"
from the one the Department uses for its own analyses. First real measurements
are July 2027 and a programme loses Direct Loan eligibility only after failing
twice in three years. **Nothing in this file has lost anything.** Any surface
built on it says so.

WHY THE CROSSWALK IS REQUIRED, and this is the finding that shaped the build:
PPD carries a `unitid` column, and it is exactly ONE UNITID PER OPEID6 -- 5,096
of each, strictly one to one. So PPD has already collapsed each Title IV
certification to a single representative campus. Joining on that unitid would
match the representative and silently miss every sibling, which for a system
like Penn State's 22 campuses is most of the institution. The join therefore
goes through opeid6, and the crosswalk decides which campuses may inherit.

WHAT IS PROPAGATED, AND WHAT IS WITHHELD. Title IV eligibility attaches to the
OPE ID rather than the campus, so propagating across a genuine multi-campus
system is correct; it is a fact about the certification and the caption must say
so, because the programme may be taught at only one campus. Where the crosswalk
says `mixed` -- different IPEDS systems, or one system whose members have
unrelated names, which is how Bard College, Simon's Rock and Longy School of
Music share an OPE ID -- nothing is emitted at all. See build_ppd_crosswalk.py.

TWO COLUMNS THAT LOOK INTERCHANGEABLE AND ARE NOT. `fail_obbb_cip2_wageb` is
the OBBBA earnings test alone. `mstr_obbb_fail_cip2_wageb` is the master flag,
which is what the published 5% figure counts. Both ship, named apart, because a
consumer quoting one under the other's headline would be wrong by a factor of
more than two: 1,220 failures against 2,880.

CREDLEV IS A STRING HERE and an integer everywhere in this repo. The map is
below and is asserted complete: an unmapped level would silently drop every
programme at that level rather than raise.

EARNINGS AND BENCHMARK ARE BOTH CARRIED, and both come from PPD. Never compare
this repo's own `earn_median` (Scorecard, different cohorts, no common
deflation) against PPD's benchmark. Both sides of the comparison come from one
file or neither does.

    python3 build_ppd_flags.py
"""
import sys
from pathlib import Path

import pandas as pd

REPO = Path(__file__).resolve().parent
EARN = REPO / "ppd2026_earnings_metrics.xlsx"
INST = REPO / "ppd2026_institutions.xlsx"
CROSSWALK = REPO / "data" / "ppd_opeid_crosswalk.csv"
OUT = REPO / "data" / "ppd_program_flags.csv"

# PPD:2026, released 2025-12-30. Stamped into the output so a consumer can date
# it, and because this file has a KNOWN EXPIRY: the July 2027 measurement
# supersedes it, after which PPD is actively misleading rather than merely
# stale.
PPD_RELEASE = "2025-12-30"
PPD_SUPERSEDED_BY = "2027-07-01"

# PPD's credlev strings against the Scorecard CREDLEV integers every other
# dataset here uses. Asserted complete at build time.
CREDLEV = {
    "Undergrad Certificate": 1, "Associate": 2, "Bachelor": 3,
    "Post-Bacc Certificate": 4, "Master's": 5, "Doctoral": 6,
    "First Professional Degree": 7, "Graduate Certificate": 8,
}

EARN_COLS = ["opeid6", "cip4", "cip4_title", "credlev", "md_earn_wne_p4",
             "count_wne_p4", "fail_obbb_cip2_wageb", "mstr_obbb_fail_cip2_wageb",
             "earn_bnchmrk_cip2_wageb", "which_test_cip2_wageb"]
INST_COLS = ["opeid6", "cip4", "credlev", "unitid"]

MIN_ROWS = 20_000          # refuse a table that has lost its shape
MIN_FAILING = 500


def read_sheet(path: Path, columns: list) -> pd.DataFrame:
    import openpyxl
    if not path.exists():
        raise SystemExit(
            f"{path.name} missing. The PPD:2026 files are a manual download from\n"
            f"  ed.gov/laws-and-policy/higher-education-laws-and-policy/"
            f"higher-education-policy/negotiated-rulemaking-for-higher-education-2025-2026\n"
            f"and are not committed. NOTE: curl gets this network's proxy block "
            f"page;\nurllib through certifi gets the real file.")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    missing = [c for c in columns if c not in header]
    if missing:
        raise SystemExit(f"{path.name} has no column(s) {missing}; PPD's schema moved")
    idx = [header.index(c) for c in columns]
    rows = [[r[i] for i in idx] for r in it]
    wb.close()
    return pd.DataFrame(rows, columns=columns)


def build() -> pd.DataFrame:
    earn = read_sheet(EARN, EARN_COLS)
    inst = read_sheet(INST, INST_COLS)

    unknown = set(earn.credlev.dropna().unique()) - set(CREDLEV)
    if unknown:
        raise SystemExit(
            f"unmapped credlev {sorted(unknown)}. An unmapped level does not "
            f"raise downstream, it silently drops every programme at that level.")

    # PPD's own unitid is one per opeid6 and is NOT used as the join key. It is
    # kept only to record which campus ED chose to represent the certification.
    ppd = earn.merge(inst, on=["opeid6", "cip4", "credlev"], how="left",
                     validate="one_to_one")
    ppd = ppd.rename(columns={"unitid": "ppd_representative_unitid"})
    ppd["CREDLEV"] = ppd.credlev.map(CREDLEV)
    ppd["CIPCODE"] = pd.to_numeric(ppd.cip4, errors="coerce")

    # Only rows with a verdict. 79% of PPD carries none, because the programme
    # is outside the statute, or privacy suppressed, or has no earnings cohort.
    # Emitting those as rows would make absence look like a measurement.
    verdict = ppd[ppd.fail_obbb_cip2_wageb.notna()
                  | ppd.mstr_obbb_fail_cip2_wageb.notna()].copy()

    cross = pd.read_csv(CROSSWALK, dtype={"OPEID6": "str"})
    allowed = cross[cross.propagate.isin(("single", "cohesive"))]
    verdict["OPEID6"] = verdict.opeid6.astype(int).astype(str).str.zfill(6)

    joined = verdict.merge(
        allowed[["OPEID6", "UNITID", "campuses", "propagate", "basis"]],
        on="OPEID6", how="inner")

    # SLIM ON PURPOSE. The first version carried cip4_title, the credential
    # string and both dates on every row, which put two CONSTANTS on 67,104
    # rows and took the file from 5.4 MB to 10.4 MB. The dates live in this
    # module and in the guard, where a reader looking for a vintage will
    # actually find them.
    out = pd.DataFrame({
        "UNITID": joined.UNITID,
        "OPEID6": joined.OPEID6,
        "CIPCODE": joined.CIPCODE,
        "CREDLEV": joined.CREDLEV,
        "obbb_fail": joined.fail_obbb_cip2_wageb,
        "master_fail": joined.mstr_obbb_fail_cip2_wageb,
        "earnings": joined.md_earn_wne_p4,
        "earnings_n": joined.count_wne_p4,
        "benchmark": joined.earn_bnchmrk_cip2_wageb,
        "benchmark_test": joined.which_test_cip2_wageb,
        "campuses": joined.campuses,
        "propagation": joined.basis,
    }).sort_values(["UNITID", "CIPCODE", "CREDLEV"]).reset_index(drop=True)
    for col in ("CIPCODE", "CREDLEV", "obbb_fail", "master_fail", "campuses"):
        out[col] = pd.to_numeric(out[col], errors="coerce").astype("Int64")

    # Printed here rather than in main(): cip4_title is deliberately not in the
    # output, so this is the last point at which the field names exist.
    titles = joined.loc[joined.mstr_obbb_fail_cip2_wageb == 1, "cip4_title"]
    print("  worst-hit fields by failing rows:")
    for title, n in titles.astype(str).str.strip().value_counts().head(5).items():
        print(f"    {n:>5}  {title[:58]}")

    failing = int((out.master_fail == 1).sum())
    if len(out) < MIN_ROWS or failing < MIN_FAILING:
        raise SystemExit(
            f"refusing to write: {len(out)} rows and {failing} failing. Below "
            f"the floors of {MIN_ROWS} and {MIN_FAILING}; the inputs or the "
            f"crosswalk probably changed shape.")
    return out


def main() -> None:
    out = build()
    print(f"{len(out):,} (school, programme) rows across "
          f"{out.UNITID.nunique():,} schools")
    for col in ("obbb_fail", "master_fail"):
        n = int((out[col] == 1).sum()); d = int(out[col].notna().sum())
        print(f"  {col:>12}: {n:,} failing of {d:,} with a verdict ({n/d:.1%})")
    print(f"  fanned beyond ED's representative campus: "
          f"{int((out.campuses > 1).sum()):,} rows")
    OUT.parent.mkdir(exist_ok=True)
    out.to_csv(OUT, index=False)
    print(f"\nwrote {OUT.relative_to(REPO)}")


if __name__ == "__main__":
    main()
